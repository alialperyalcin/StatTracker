from __future__ import annotations

import os
import time
from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook, load_workbook

from stat_extractor import CANONICAL_FIELDS


HEADER: List[str] = ["Timestamp", "Image File", *CANONICAL_FIELDS]


class ExcelFileLockedError(RuntimeError):
    pass


def _ensure_workbook(path: str):
    try:
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
            if ws.max_row < 1:
                ws.append(HEADER)
            else:
                first_row = [cell.value for cell in ws[1]]
                if first_row != HEADER:
                    ws.insert_rows(1)
                    for idx, name in enumerate(HEADER, start=1):
                        ws.cell(row=1, column=idx, value=name)
            return wb, ws

        wb = Workbook()
        ws = wb.active
        ws.title = "Stats"
        ws.append(HEADER)
        return wb, ws
    except PermissionError as exc:
        raise ExcelFileLockedError(
            f"Excel file is locked: '{path}'. Close it in Excel/OneDrive and try again."
        ) from exc


def _save_with_retry(wb: Workbook, path: str, retries: int = 5, wait_sec: float = 0.4) -> None:
    last_exc: Exception | None = None
    for _ in range(retries):
        try:
            wb.save(path)
            return
        except PermissionError as exc:
            last_exc = exc
            time.sleep(wait_sec)

    raise ExcelFileLockedError(
        f"Could not write to '{path}' because it is locked by another app."
    ) from last_exc


def append_stats(path: str, stats: Dict[str, int], image_file: str) -> None:
    wb, ws = _ensure_workbook(path)

    row = [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), os.path.basename(image_file)]
    for key in CANONICAL_FIELDS:
        row.append(stats.get(key))
    ws.append(row)
    try:
        _save_with_retry(wb, path)
    except PermissionError as exc:
        raise ExcelFileLockedError(
            f"Excel file is locked: '{path}'. Close it in Excel/OneDrive and try again."
        ) from exc
